from codexcost.core import TokenCount
import logging
from pathlib import Path
from typing import Iterable, Any

try:
    import openpyxl
    import openpyxl.utils as oputils
    import openpyxl.worksheet.table as optable
except ModuleNotFoundError as e:
    logging.error('Unable to find openpyxl. Please install openpyxl>=3.1.5', exc_info=e)
    raise


def write_xlsx(token_counts: Iterable[TokenCount], file:Path):
    '''Write list of TokenCount to xlsx'''
    logging.info('Writing token counts to xlsx. file=%s', file)
    
    def transform(token_count:TokenCount) -> dict[str,Any]:
        return {'Timestamp': token_count.timestamp.replace(tzinfo=None),
            'Session': token_count.id, 'Model': token_count.model, 'Project': token_count.project,
            'Uncached Input': token_count.uncached_input_tokens, 'Cached Input': token_count.cached_input_tokens,
            'Output': token_count.output_tokens, 'Credits': token_count.credits}

    headers = (('Timestamp', 'datetime'), ('Session', 'str'), ('Model', 'str'), ('Project', 'str'),
        ('Uncached Input', 'int'), ('Cached Input', 'int'), ('Output', 'int'), ('Credits', 'float'))

    __write_excel_worksheet_table(map(transform, token_counts), headers, file, sheet_name='Codex Token Usage')

__EXCEL_NUMBER_FORMAT: dict[str,str] = {
    'int': '0',
    'float': '0.00',
    'datetime': 'yyyy-mm-dd hh:mm:ss',
    'date': 'yyyy-mm-dd',
    'bool': '',
}

def __write_excel_worksheet_table(data:Iterable[dict], headers:Iterable[tuple[str, str]], export_file:Path, sheet_name:str='Sheet1',
        _table_style='TableStyleMedium1'):
    '''Write tabular data into an excel table.
    '''
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    # write header
    ws.append(header[0] for header in headers)

    # write row
    for ri, row in enumerate(data, start=2):
        for ci, (column, typ) in enumerate(headers, start=1):
            cell = ws.cell(row=ri, column=ci, value=row[column])
            cell.number_format = __EXCEL_NUMBER_FORMAT.get(typ, typ)

    # write table
    table = optable.Table(displayName='Table1', ref=f'A1:{oputils.get_column_letter(ws.max_column)}{ws.max_row}')
    table.tableStyleInfo = optable.TableStyleInfo(name=_table_style,
        showFirstColumn=True, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False)
    ws.add_table(table)

    wb.save(export_file)
